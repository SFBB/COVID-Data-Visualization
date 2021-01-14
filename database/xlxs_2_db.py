import openpyxl
import pandas as pd
from pathlib import Path
import json
import hashlib
import psycopg2
from django.core.serializers.json import DjangoJSONEncoder


def connect_db(name):
    try:
        con = psycopg2.connect(database=name, user="username", password="password", host="127.0.0.1", port="5432")
        con.autocommit = True

        # sqliteConnection = sqlite3.connect(name)
        print("Successfully Connected to SQLite")
        # print("Successfully Connected to SQLite")

        return con
    except Exception as error:
        print("Failed to insert data into sqlite table "+str(error))
        # print("Failed to insert data into sqlite table", error)


def insert_to_db(conn, table, ids, values):
    try:
        sql = '''  INSERT INTO '''+table+'''('''+ids+''') 
                VALUES('''+values+'''); '''

        cur = conn.cursor()
        cur.execute(sql)
        conn.commit()
        cur.close()
        return cur.lastrowid
    except Exception as error:
        print("Failed to insert data into sqlite table "+str(error)+"\t\t\t"+sql)
        # print("Failed to insert data into sqlite table", error)
        # print(sql)
        # print()
        # print()
        # print()


def in_db(conn, table, id):
    try:
        # conn = sqlite3.connect("narou.db")
        sql = '''  SELECT * FROM '''+table+'''
                    WHERE id='''+id+''' '''
        cur = conn.cursor()
        cur.execute(sql)
        conn.commit()
        result = cur.fetchall()
        cur.close()
        if len(result) > 0:
            return True
        else:
            return False
    except Exception as error:
        print("Failed to get data into sqlite table "+str(error)+"\t\t\t"+sql)
        # print("Failed to get data into sqlite table", error)
        # print(sql)
        # print()
        # print()
        # print()




def read_data(sheet, db):
    # print(sheet.tables)
    # print(sheet)
    # print(sheet["C2"].value)
    count = 0
    data = {}
    keys = []
    ids = """id, 国家地区, 日期, 累计确诊, 新增确诊, 累计死亡, 新增死亡, 累计治愈, 仍在治疗, 重症病例, 百万人口确诊率, 百万人口死亡率, 总检测数, 百万人口检测率"""
    # for column in sheet.iter_cols(1, sheet.max_column, values_only=True):
    #     if column[0] == None:
    #         break
    #     if count < 1:
    #         for row in column:
    #             if row not in data:
    #                 data[row] = {"data": [], "numbers": 0}
    #             # else:
    #                 # data[row]["data"].append()
    #             # print(row)
    #     keys.append(column[0])
    #     # else:
    #         # template[column[0]] = ""
    #     count += 1
    #     # print(column)
    #     # if
    #     # print(column[0])

    # print([key for key in template.keys()][0])

    # keys = [key for key in template.keys()]
    for row in sheet.iter_rows(2, sheet.max_row, values_only=True):
        if row[0] == None:
            break
        temp = []
        values = ""

        # print(str(row[1]))
        md5 = hashlib.md5()
        md5.update((row[0]+str(row[1])).encode('utf-8'))
        id = "'"+md5.hexdigest()+"'"
        # print(id)
        values = id+", "
        for i in range(1):
            # pass
            values += "'"+row[i]+"', "
        for i in range(1, 2):
            # pass
            values += "'"+str(row[i])+"', "
        for i in range(2, 13):
            if row[i] == None or len(str(row[i]))==0 or str(row[i])=="N/A":
                values += "NULL, "
            else:
                values += str(row[i])+", "
        values = values[:len(values)-2]
        # print(values)
        if not in_db(db, "yiqing", id):
            insert_to_db(db, "yiqing", ids, values)
        # print(values)
        # break
        # data[row[0]]["data"].append(temp)
            # print(row[i])
        # if count < 1:
        #     for row in column:
        #         if row not in data:
        #             data[row] = {"data": [], "numbers": 0}
        #         # else:
        #             # data[row]["data"].append()
        #         # print(row)
        # else:
        #     template[column[0]] = ""
        # count += 1
        # print(column)
        # if
        # print(column[0])

    # print(data["全球"]["data"][0])
    # return data






if __name__=="__main__":
    db = connect_db("yiqing")


    xlsx_file = Path('data', '3.xlsx')
    # df = pd.read_excel ("data/3.xlsx", engine="openpyxl")
    # print(df)
    wb_obj = openpyxl.load_workbook(xlsx_file) 

    # # Read the active sheet:
    sheets = wb_obj.sheetnames
    # print(wb_obj[sheets[1]])
    for i in range(len(sheets)):
        read_data(wb_obj[sheets[i]], db)

    # data = read_data(sheet)

    # with open("data/全球及重点国家疫情主要指数数据-2020-6-22-7H-1.json", "w") as file:
    #     json.dump(data, file, cls=DjangoJSONEncoder)



    # xlsx_file = Path('data', '全球及重点国家疫情主要指数数据-2020-6-29-7H.xlsx')
    # wb_obj = openpyxl.load_workbook(xlsx_file) 

    # # Read the active sheet:
    # for sheet in wb_obj:
    #     print(sheet)
    #     sheet = wb_obj.active

    #     data = read_data(sheet)

    #     # with open("data/全球及重点国家疫情主要指数数据-2020-6-29-7H.json", "w") as file:
    #     #     json.dump(data, file, cls=DjangoJSONEncoder)
